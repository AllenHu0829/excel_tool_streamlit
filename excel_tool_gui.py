import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import shutil
import threading
from PIL import Image, ImageTk


class ExcelToolGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel文件拆分与合并工具")
        self.root.geometry("850x750")
        
        # iOS风格颜色主题
        self.colors = {
            'bg': '#f2f2f7',  # iOS系统背景色
            'primary': '#007AFF',  # iOS蓝色
            'primary_hover': '#0051D5',
            'success': '#34C759',  # iOS绿色
            'success_hover': '#30D158',
            'card_bg': '#ffffff',
            'card_shadow': '#e5e5ea',
            'text': '#000000',
            'text_secondary': '#8e8e93',  # iOS次要文字色
            'separator': '#c6c6c8',
            'header_bg': '#ffffff',
            'button_text': '#ffffff',
            'section_bg': '#f9f9f9'
        }
        
        # iOS风格字体
        self.fonts = {
            'title': ('PingFang SC', 28, 'normal'),  # 标题字体，更大
            'subtitle': ('PingFang SC', 18, 'normal'),
            'body': ('PingFang SC', 16, 'normal'),  # 正文字体
            'body_small': ('PingFang SC', 14, 'normal'),
            'button': ('PingFang SC', 17, 'normal'),  # 按钮字体
            'label': ('PingFang SC', 15, 'normal'),
            'mono': ('SF Mono', 13, 'normal')  # 等宽字体用于路径
        }
        
        # 如果没有PingFang SC，使用备用字体
        try:
            # 测试字体是否存在
            test_label = tk.Label(self.root, text="测试", font=self.fonts['title'])
            test_label.destroy()
        except:
            # 使用Windows系统字体作为备用
            self.fonts = {
                'title': ('Segoe UI', 24, 'normal'),
                'subtitle': ('Segoe UI', 18, 'normal'),
                'body': ('Segoe UI', 15, 'normal'),
                'body_small': ('Segoe UI', 13, 'normal'),
                'button': ('Segoe UI', 15, 'normal'),
                'label': ('Segoe UI', 14, 'normal'),
                'mono': ('Consolas', 12, 'normal')
            }
        
        self.root.configure(bg=self.colors['bg'])
        
        # 变量
        self.mode = tk.StringVar(value="split")  # split 或 merge
        self.source_path = tk.StringVar()
        self.output_path = tk.StringVar()
        self.execute_btn = None
        
        self.create_widgets()
        
    def create_widgets(self):
        # 主容器 - iOS风格的大间距
        main_container = tk.Frame(self.root, bg=self.colors['bg'])
        main_container.pack(fill="both", expand=True, padx=20, pady=25)
        
        # iOS风格标题区域（带背景图片）
        header_frame = tk.Frame(main_container, bg=self.colors['card_bg'], 
                              relief="flat", bd=0, height=140)
        header_frame.pack(fill="x", pady=(0, 30))
        header_frame.pack_propagate(False)
        
        # 创建Canvas用于显示背景图片
        self.header_canvas = tk.Canvas(header_frame, 
                                      bg=self.colors['card_bg'],
                                      highlightthickness=0,
                                      relief="flat",
                                      bd=0,
                                      width=850,
                                      height=140)
        self.header_canvas.pack(fill="both", expand=True)
        
        # 尝试加载背景图片
        self.header_bg_image = None
        has_image = self.load_header_background()
        
        # 如果图片加载成功，在Canvas上绘制
        if self.header_bg_image and has_image:
            self.header_canvas.create_image(0, 0, anchor="nw", image=self.header_bg_image)
            # 添加半透明遮罩，使文字更清晰
            self.header_canvas.create_rectangle(0, 0, 850, 140, 
                                               fill="#ffffff", 
                                               stipple="gray25",
                                               outline="")
        
        # 创建文字层（在背景图片之上）- 使用Canvas的create_text而不是Frame
        title_label = self.header_canvas.create_text(30, 50, 
                              text="Excel 工具", 
                              font=self.fonts['title'],
                              fill=self.colors['text'],
                              anchor="w")
        
        subtitle_label = self.header_canvas.create_text(30, 85, 
                                 text="拆分与合并 Excel 文件", 
                                 font=self.fonts['subtitle'],
                                 fill=self.colors['text_secondary'],
                                 anchor="w")
        
        # iOS风格功能选择区域（卡片式设计）
        mode_frame = self.create_ios_card(main_container)
        mode_frame.pack(pady=(0, 15), fill="x")
        
        mode_title = tk.Label(mode_frame, text="选择功能", 
                             font=self.fonts['body_small'],
                             bg=self.colors['card_bg'], 
                             fg=self.colors['text_secondary'],
                             anchor="w")
        mode_title.pack(fill="x", padx=20, pady=(20, 10))
        
        # 分隔线（iOS风格）
        separator1 = tk.Frame(mode_frame, bg=self.colors['separator'], height=1)
        separator1.pack(fill="x", padx=20)
        
        mode_inner = tk.Frame(mode_frame, bg=self.colors['card_bg'])
        mode_inner.pack(fill="x", padx=20, pady=15)
        
        split_radio = tk.Radiobutton(mode_inner, text="拆分 Excel 文件", 
                                     variable=self.mode, value="split",
                                     command=self.on_mode_change,
                                     font=self.fonts['body'],
                                     bg=self.colors['card_bg'], 
                                     fg=self.colors['text'],
                                     selectcolor=self.colors['card_bg'],
                                     activebackground=self.colors['card_bg'],
                                     activeforeground=self.colors['primary'],
                                     cursor="hand2",
                                     indicatoron=True)
        split_radio.pack(anchor="w", pady=8)
        
        merge_radio = tk.Radiobutton(mode_inner, text="合并 Excel 文件", 
                                    variable=self.mode, value="merge",
                                    command=self.on_mode_change,
                                    font=self.fonts['body'],
                                    bg=self.colors['card_bg'], 
                                    fg=self.colors['text'],
                                    selectcolor=self.colors['card_bg'],
                                    activebackground=self.colors['card_bg'],
                                    activeforeground=self.colors['primary'],
                                    cursor="hand2",
                                    indicatoron=True)
        merge_radio.pack(anchor="w", pady=8)
        
        # iOS风格源文件/文件夹选择区域
        source_frame = self.create_ios_card(main_container)
        source_frame.pack(pady=(0, 15), fill="x")
        
        self.source_label_text = tk.StringVar(value="源文件")
        source_title = tk.Label(source_frame, textvariable=self.source_label_text,
                               font=self.fonts['body_small'],
                               bg=self.colors['card_bg'], 
                               fg=self.colors['text_secondary'],
                               anchor="w")
        source_title.pack(fill="x", padx=20, pady=(20, 10))
        
        separator2 = tk.Frame(source_frame, bg=self.colors['separator'], height=1)
        separator2.pack(fill="x", padx=20)
        
        source_input_frame = tk.Frame(source_frame, bg=self.colors['card_bg'])
        source_input_frame.pack(fill="x", padx=20, pady=15)
        
        # iOS风格输入框（更大的内边距）
        entry_frame = tk.Frame(source_input_frame, bg=self.colors['section_bg'],
                               relief="flat", bd=0)
        entry_frame.pack(side="left", fill="x", expand=True, ipady=12, ipadx=15)
        
        self.source_entry = tk.Label(entry_frame, textvariable=self.source_path,
                                     font=self.fonts['mono'],
                                     bg=self.colors['section_bg'], 
                                     fg=self.colors['text'],
                                     anchor="w",
                                     wraplength=400)
        self.source_entry.pack(fill="both", expand=True, padx=5)
        
        source_btn = self.create_ios_button(source_input_frame, "选择", 
                                            self.browse_source,
                                            bg=self.colors['primary'],
                                            hover_bg=self.colors['primary_hover'],
                                            font=self.fonts['button'],
                                            padx=20, pady=12)
        source_btn.pack(side="right", padx=(12, 0))
        
        # iOS风格输出路径选择区域
        output_frame = self.create_ios_card(main_container)
        output_frame.pack(pady=(0, 20), fill="x")
        
        self.output_label_text = tk.StringVar(value="输出路径")
        output_title = tk.Label(output_frame, textvariable=self.output_label_text,
                               font=self.fonts['body_small'],
                               bg=self.colors['card_bg'], 
                               fg=self.colors['text_secondary'],
                               anchor="w")
        output_title.pack(fill="x", padx=20, pady=(20, 10))
        
        separator3 = tk.Frame(output_frame, bg=self.colors['separator'], height=1)
        separator3.pack(fill="x", padx=20)
        
        output_input_frame = tk.Frame(output_frame, bg=self.colors['card_bg'])
        output_input_frame.pack(fill="x", padx=20, pady=15)
        
        entry_frame2 = tk.Frame(output_input_frame, bg=self.colors['section_bg'],
                               relief="flat", bd=0)
        entry_frame2.pack(side="left", fill="x", expand=True, ipady=12, ipadx=15)
        
        self.output_entry = tk.Label(entry_frame2, textvariable=self.output_path,
                                     font=self.fonts['mono'],
                                     bg=self.colors['section_bg'], 
                                     fg=self.colors['text'],
                                     anchor="w",
                                     wraplength=400)
        self.output_entry.pack(fill="both", expand=True, padx=5)
        
        output_btn = self.create_ios_button(output_input_frame, "选择", 
                                           self.browse_output,
                                           bg=self.colors['primary'],
                                           hover_bg=self.colors['primary_hover'],
                                           font=self.fonts['button'],
                                           padx=20, pady=12)
        output_btn.pack(side="right", padx=(12, 0))
        
        # iOS风格执行按钮（大按钮，全宽）
        button_frame = tk.Frame(main_container, bg=self.colors['bg'])
        button_frame.pack(pady=(10, 20), fill="x")
        
        self.execute_btn = self.create_ios_button(button_frame, "开始执行", 
                                                  self.execute_task,
                                                  bg=self.colors['success'],
                                                  hover_bg=self.colors['success_hover'],
                                                  font=self.fonts['button'],
                                                  padx=0, pady=16,
                                                  full_width=True)
        
        # iOS风格状态显示区域
        status_frame = self.create_ios_card(main_container)
        status_frame.pack(pady=(0, 15), fill="both", expand=True)
        
        status_title = tk.Label(status_frame, text="执行状态", 
                               font=self.fonts['body_small'],
                               bg=self.colors['card_bg'], 
                               fg=self.colors['text_secondary'],
                               anchor="w")
        status_title.pack(fill="x", padx=20, pady=(20, 10))
        
        separator4 = tk.Frame(status_frame, bg=self.colors['separator'], height=1)
        separator4.pack(fill="x", padx=20)
        
        # 使用Text组件和Scrollbar显示状态
        text_frame = tk.Frame(status_frame, bg=self.colors['card_bg'])
        text_frame.pack(fill="both", expand=True, padx=20, pady=15)
        
        scrollbar = tk.Scrollbar(text_frame, 
                                troughcolor=self.colors['bg'],
                                activebackground=self.colors['text_secondary'],
                                width=12,
                                relief="flat",
                                bd=0)
        scrollbar.pack(side="right", fill="y", pady=2)
        
        self.status_text = tk.Text(text_frame, 
                                  yscrollcommand=scrollbar.set,
                                  wrap=tk.WORD, state="disabled",
                                  font=self.fonts['body_small'],
                                  bg=self.colors['card_bg'], 
                                  fg=self.colors['text'],
                                  relief="flat", bd=0,
                                  padx=0, pady=5,
                                  selectbackground=self.colors['primary'],
                                  selectforeground="white",
                                  spacing1=2,
                                  spacing2=2,
                                  spacing3=2)
        self.status_text.pack(side="left", fill="both", expand=True)
        scrollbar.config(command=self.status_text.yview)
        
        # 配置文本标签颜色（在第一次使用前）
        self.status_text.tag_config("error", foreground="#FF3B30")
        self.status_text.tag_config("success", foreground="#34C759")
        self.status_text.tag_config("warning", foreground="#FF9500")
        self.status_text.tag_config("normal", foreground=self.colors['text'])
        
        # 初始化
        self.on_mode_change()
        
    def load_header_background(self):
        """加载标题区域的背景图片，按比例缩小并靠右摆放"""
        # 目标尺寸
        target_width = 850
        target_height = 140
        
        # 可能的图片路径
        possible_paths = [
            os.path.join(os.path.dirname(__file__), "cat_bg.png"),
            os.path.join(os.path.dirname(__file__), "cat_bg.jpg"),
            os.path.join(os.path.dirname(__file__), "header_bg.png"),
            os.path.join(os.path.dirname(__file__), "header_bg.jpg"),
            "cat_bg.png",
            "cat_bg.jpg",
        ]
        
        image_path = None
        for path in possible_paths:
            if os.path.exists(path):
                image_path = path
                break
        
        if image_path:
            try:
                # 打开原始图片
                original_img = Image.open(image_path)
                orig_width, orig_height = original_img.size
                
                # 计算缩放比例（缩小到目标区域内，保持宽高比）
                # 选择较小的缩放比例，确保图片完全在目标区域内
                scale_width = target_width / orig_width
                scale_height = target_height / orig_height
                scale = min(scale_width, scale_height)
                
                # 如果图片已经小于目标尺寸，可以选择不放大或保持原尺寸
                # 这里设置为最多缩小到目标区域，如果更小则保持原尺寸
                if scale > 1.0:
                    # 如果图片比目标区域小，可以选择保持原尺寸或放大
                    # 这里选择不放大，保持原尺寸
                    scale = 1.0
                
                # 计算缩放后的尺寸
                new_width = int(orig_width * scale)
                new_height = int(orig_height * scale)
                
                # 高质量缩放
                img = original_img.resize((new_width, new_height), Image.Resampling.LANCZOS)
                
                # 创建目标尺寸的画布（白色背景）
                canvas_img = Image.new('RGB', (target_width, target_height), color='#ffffff')
                
                # 计算靠右对齐位置
                paste_x = target_width - new_width  # 靠右
                paste_y = (target_height - new_height) // 2  # 垂直居中
                
                # 确保不会超出边界
                if paste_x < 0:
                    paste_x = 0
                if paste_y < 0:
                    paste_y = 0
                
                # 将缩放后的图片粘贴到画布右侧
                canvas_img.paste(img, (paste_x, paste_y))
                
                self.header_bg_image = ImageTk.PhotoImage(canvas_img)
                return True
            except Exception as e:
                print(f"加载背景图片失败: {e}")
                self.header_bg_image = None
                return False
        else:
            # 如果没有找到图片，创建一个默认背景
            try:
                img = Image.new('RGB', (target_width, target_height), color='#ffffff')
                self.header_bg_image = ImageTk.PhotoImage(img)
                return False  # 返回False表示使用默认背景
            except:
                self.header_bg_image = None
                return False
        
    def create_ios_card(self, parent):
        """创建iOS风格的卡片"""
        card = tk.Frame(parent, bg=self.colors['card_bg'], 
                       relief="flat", bd=0)
        return card
        
    def create_ios_button(self, parent, text, command, bg="#007AFF", hover_bg="#0051D5",
                         fg="white", font=None, padx=20, pady=14, full_width=False):
        """创建iOS风格的按钮"""
        if font is None:
            font = self.fonts['button']
            
        btn = tk.Button(parent, text=text, command=command,
                       bg=bg, fg=fg, font=font,
                       padx=padx, pady=pady,
                       relief="flat", bd=0,
                       cursor="hand2",
                       activebackground=hover_bg,
                       activeforeground=fg,
                       highlightthickness=0,
                       borderwidth=0)
        
        if full_width:
            btn.pack(fill="x")
        else:
            btn.pack()
        
        # 添加悬停效果
        def on_enter(e):
            btn.config(bg=hover_bg)
        
        def on_leave(e):
            btn.config(bg=bg)
        
        btn.bind("<Enter>", on_enter)
        btn.bind("<Leave>", on_leave)
        
        return btn
        
        
    def on_mode_change(self):
        """当模式改变时更新界面"""
        mode = self.mode.get()
        if mode == "split":
            self.source_label_text.set("源文件")
            self.output_label_text.set("输出路径")
            self.source_path.set("")
            self.output_path.set("")
        else:
            self.source_label_text.set("源文件夹")
            self.output_label_text.set("输出文件")
            self.source_path.set("")
            self.output_path.set("")
        self.log_message(f"模式已切换: {'拆分' if mode == 'split' else '合并'}")
        
    def browse_source(self):
        """浏览源文件/文件夹"""
        mode = self.mode.get()
        if mode == "split":
            filename = filedialog.askopenfilename(
                title="选择要拆分的Excel文件",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            )
            if filename:
                self.source_path.set(filename)
                self.log_message(f"已选择源文件: {filename}")
        else:
            dirname = filedialog.askdirectory(title="选择包含Excel文件的文件夹")
            if dirname:
                self.source_path.set(dirname)
                self.log_message(f"已选择源文件夹: {dirname}")
                
    def browse_output(self):
        """浏览输出路径"""
        mode = self.mode.get()
        if mode == "split":
            dirname = filedialog.askdirectory(title="选择拆分文件保存文件夹")
            if dirname:
                self.output_path.set(dirname)
                self.log_message(f"已选择输出文件夹: {dirname}")
        else:
            filename = filedialog.asksaveasfilename(
                title="选择合并后文件保存路径",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            if filename:
                self.output_path.set(filename)
                self.log_message(f"已选择输出文件: {filename}")
                
    def log_message(self, message):
        """在状态区域显示消息"""
        self.status_text.config(state="normal")
        # 根据消息类型设置不同的颜色
        tag = "normal"
        if "错误" in message or "失败" in message:
            tag = "error"
        elif "完成" in message or "成功" in message:
            tag = "success"
        elif "警告" in message:
            tag = "warning"
        
        self.status_text.insert(tk.END, message + "\n", tag)
        self.status_text.see(tk.END)
        self.status_text.config(state="disabled")
        self.root.update()
        
    def execute_task(self):
        """执行拆分或合并任务"""
        source = self.source_path.get()
        output = self.output_path.get()
        mode = self.mode.get()
        
        # 验证输入
        if not source:
            messagebox.showerror("错误", "请选择源文件/文件夹！")
            return
            
        if not output:
            messagebox.showerror("错误", "请选择输出路径！")
            return
            
        if mode == "split" and not os.path.isfile(source):
            messagebox.showerror("错误", "源文件不存在！")
            return
            
        if mode == "merge" and not os.path.isdir(source):
            messagebox.showerror("错误", "源文件夹不存在！")
            return
        
        # 在新线程中执行任务，避免界面卡顿
        thread = threading.Thread(target=self.run_task, args=(mode, source, output))
        thread.daemon = True
        thread.start()
        
    def run_task(self, mode, source, output):
        """在后台线程中运行任务"""
        try:
            if mode == "split":
                self.log_message("=" * 50)
                self.log_message("开始拆分Excel文件...")
                self.split_excel_by_rows(source, output)
                self.log_message("拆分完成！")
                messagebox.showinfo("成功", "文件拆分完成！")
            else:
                self.log_message("=" * 50)
                self.log_message("开始合并Excel文件...")
                self.merge_excel_files(source, output)
                self.log_message("合并完成！")
                messagebox.showinfo("成功", "文件合并完成！")
        except Exception as e:
            error_msg = f"执行过程中出错: {str(e)}"
            self.log_message(error_msg)
            messagebox.showerror("错误", error_msg)
            
    def split_excel_by_rows(self, input_file, output_dir):
        """按照表头分割Excel文件，每一行对应一个文件"""
        try:
            # 使用openpyxl读取原始文件
            self.log_message(f"正在读取文件: {input_file}")
            source_wb = load_workbook(input_file)
            source_ws = source_wb.active
            
            self.log_message(f"Excel文件结构: 最大行数={source_ws.max_row}, 最大列数={source_ws.max_column}")
            
            # 创建输出目录
            if os.path.exists(output_dir):
                try:
                    shutil.rmtree(output_dir)
                    self.log_message("已清理旧的输出文件夹")
                except PermissionError:
                    self.log_message("警告: 无法删除旧文件，将覆盖现有文件")
            os.makedirs(output_dir, exist_ok=True)
            
            # 定义颜色填充
            blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            
            # 遍历每一行数据（从第2行开始，因为第1行是表头）
            file_count = 0
            for row_num in range(2, source_ws.max_row + 1):
                # 检查该行是否有数据（检查A列是否有内容）
                if source_ws.cell(row=row_num, column=1).value is None:
                    continue
                
                # 创建新的工作簿
                wb = Workbook()
                ws = wb.active
                
                # 复制表头第1行
                for col in range(1, source_ws.max_column + 1):
                    source_cell = source_ws.cell(row=1, column=col)
                    target_cell = ws.cell(row=1, column=col)
                    target_cell.value = source_cell.value
                    
                    # 应用颜色填充
                    if 6 <= col <= 11:  # F1~K1 (列6-11)
                        target_cell.fill = blue_fill
                    elif 12 <= col <= 13:  # L1~M1 (列12-13)
                        target_cell.fill = red_fill
                
                # 复制数据行（第2行）
                for col in range(1, source_ws.max_column + 1):
                    source_cell = source_ws.cell(row=row_num, column=col)
                    target_cell = ws.cell(row=2, column=col)
                    target_cell.value = source_cell.value
                
                # 自动调整列宽
                for col in range(1, source_ws.max_column + 1):
                    column_letter = ws.cell(row=1, column=col).column_letter
                    max_length = 0
                    
                    # 检查表头和数据行的内容长度
                    for row in range(1, 3):
                        cell_value = ws.cell(row=row, column=col).value
                        if cell_value:
                            # 计算字符长度，中文字符按2个字符计算
                            length = 0
                            for char in str(cell_value):
                                if ord(char) > 127:
                                    length += 2
                                else:
                                    length += 1
                            max_length = max(max_length, length)
                    
                    # 设置列宽，最小宽度为8，最大宽度为50
                    adjusted_width = min(max(max_length + 2, 8), 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # 获取该文件A2单元格的内容作为文件名
                filename_base = str(ws.cell(row=2, column=1).value) if ws.cell(row=2, column=1).value else f"file_{file_count + 1}"
                
                # 清理文件名中的非法字符
                filename_base = "".join(c for c in filename_base if c.isalnum() or c in (' ', '-', '_', '(', ')', '（', '）', '，', '。')).strip()
                if not filename_base:
                    filename_base = f"file_{file_count + 1}"
                
                # 生成文件名
                filename = f"{filename_base}.xlsx"
                output_path = os.path.join(output_dir, filename)
                
                # 如果文件名已存在，添加序号
                counter = 1
                original_filename = filename
                while os.path.exists(output_path):
                    name, ext = os.path.splitext(original_filename)
                    filename = f"{name}_{counter}{ext}"
                    output_path = os.path.join(output_dir, filename)
                    counter += 1
                
                # 保存文件
                wb.save(output_path)
                file_count += 1
                if file_count % 10 == 0:  # 每10个文件输出一次进度
                    self.log_message(f"已创建 {file_count} 个文件...")
            
            self.log_message(f"\n分割完成！共创建了 {file_count} 个文件")
            self.log_message(f"文件保存在: {output_dir}")
            
        except Exception as e:
            self.log_message(f"处理文件时出错: {str(e)}")
            raise
            
    def merge_excel_files(self, data_dir, output_file):
        """合并指定文件夹下的所有 Excel 文件"""
        try:
            # 获取所有 Excel 文件
            excel_files = []
            for file in os.listdir(data_dir):
                if file.endswith('.xlsx') or file.endswith('.xls'):
                    excel_files.append(os.path.join(data_dir, file))
            
            if not excel_files:
                raise Exception("文件夹下没有找到 Excel 文件")
            
            self.log_message(f"找到 {len(excel_files)} 个 Excel 文件")
            
            # 存储所有数据框
            dataframes = []
            
            # 读取每个 Excel 文件
            for idx, file_path in enumerate(excel_files, 1):
                try:
                    # 读取 Excel 文件，使用第一行作为列名
                    df = pd.read_excel(file_path, header=0)
                    
                    # 添加源文件名列，用于追踪数据来源
                    if '源文件' not in df.columns:
                        df.insert(0, '源文件', os.path.basename(file_path))
                    
                    dataframes.append(df)
                    self.log_message(f"已读取 [{idx}/{len(excel_files)}]: {os.path.basename(file_path)} - {df.shape[0]} 行, {df.shape[1]} 列")
                    
                except Exception as e:
                    self.log_message(f"读取文件失败 {os.path.basename(file_path)}: {str(e)}")
                    continue
            
            if not dataframes:
                raise Exception("没有成功读取任何文件")
            
            # 合并所有数据框
            self.log_message("\n正在合并数据...")
            merged_df = pd.concat(dataframes, ignore_index=True, sort=False)
            
            # 统计信息
            self.log_message(f"\n合并统计:")
            self.log_message(f"总行数: {len(merged_df)}")
            self.log_message(f"总列数: {len(merged_df.columns)}")
            self.log_message(f"列名: {list(merged_df.columns)}")
            
            # 保存合并后的文件
            self.log_message(f"\n正在保存到: {output_file}")
            merged_df.to_excel(output_file, index=False, engine='openpyxl')
            self.log_message("保存完成！")
            
        except Exception as e:
            self.log_message(f"处理过程中出错: {str(e)}")
            raise


def main():
    root = tk.Tk()
    app = ExcelToolGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()

