from openpyxl import load_workbook

def check_excel_structure(input_file):
    """检查Excel文件的结构"""
    try:
        wb = load_workbook(input_file)
        ws = wb.active
        
        print(f"Excel文件结构:")
        print(f"最大行数: {ws.max_row}")
        print(f"最大列数: {ws.max_column}")
        print()
        
        # 检查前几行的内容
        for row in range(1, min(6, ws.max_row + 1)):
            print(f"第{row}行内容:")
            for col in range(1, min(20, ws.max_column + 1)):  # 显示前19列
                cell_value = ws.cell(row=row, column=col).value
                print(f"  {chr(64+col)}{row}: {cell_value}")
            print()
        
        # 特别检查A2单元格内容
        a2_value = ws.cell(row=2, column=1).value
        print(f"A2单元格内容: {a2_value}")
        
        # 检查合并单元格
        if ws.merged_cells:
            print(f"\n合并单元格:")
            for merged_range in ws.merged_cells:
                print(f"  {merged_range}")
        
    except Exception as e:
        print(f"检查文件时出错: {str(e)}")

if __name__ == "__main__":
    input_file = r"c:\Users\AllenHu\excel data\工作簿1.xlsx"
    check_excel_structure(input_file)

