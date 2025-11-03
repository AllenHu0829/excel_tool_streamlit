"""
Excel文件拆分与合并工具 - Streamlit版本
"""

import streamlit as st
import pandas as pd
import os
import tempfile
import zipfile
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import shutil
import io


def split_excel_by_rows(input_file, output_dir):
    """按照表头分割Excel文件，每一行对应一个文件"""
    try:
        # 使用openpyxl读取原始文件
        source_wb = load_workbook(input_file)
        source_ws = source_wb.active
        
        # 创建输出目录
        if os.path.exists(output_dir):
            try:
                shutil.rmtree(output_dir)
            except PermissionError:
                pass
        os.makedirs(output_dir, exist_ok=True)
        
        # 定义颜色填充
        blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        
        # 进度条
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        # 遍历每一行数据（从第2行开始，因为第1行是表头）
        file_count = 0
        total_rows = source_ws.max_row - 1  # 排除表头行
        
        for row_num in range(2, source_ws.max_row + 1):
            # 检查该行是否有数据（检查A列是否有内容）
            if source_ws.cell(row=row_num, column=1).value is None:
                continue
            
            # 创建新的工作簿
            wb = Workbook()
            ws = wb.active
            
            # 复制表头第1行
            for col in range(1, source_ws.max_column + 1):
                source_cell = source_ws.cell(row=1, column=col)
                target_cell = ws.cell(row=1, column=col)
                target_cell.value = source_cell.value
                
                # 应用颜色填充
                if 6 <= col <= 11:  # F1~K1 (列6-11)
                    target_cell.fill = blue_fill
                elif 12 <= col <= 13:  # L1~M1 (列12-13)
                    target_cell.fill = red_fill
            
            # 复制数据行（第2行）
            for col in range(1, source_ws.max_column + 1):
                source_cell = source_ws.cell(row=row_num, column=col)
                target_cell = ws.cell(row=2, column=col)
                target_cell.value = source_cell.value
            
            # 自动调整列宽
            for col in range(1, source_ws.max_column + 1):
                column_letter = ws.cell(row=1, column=col).column_letter
                max_length = 0
                
                # 检查表头和数据行的内容长度
                for row in range(1, 3):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        # 计算字符长度，中文字符按2个字符计算
                        length = 0
                        for char in str(cell_value):
                            if ord(char) > 127:
                                length += 2
                            else:
                                length += 1
                        max_length = max(max_length, length)
                
                # 设置列宽，最小宽度为8，最大宽度为50
                adjusted_width = min(max(max_length + 2, 8), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 获取该文件A2单元格的内容作为文件名
            filename_base = str(ws.cell(row=2, column=1).value) if ws.cell(row=2, column=1).value else f"file_{file_count + 1}"
            
            # 清理文件名中的非法字符
            filename_base = "".join(c for c in filename_base if c.isalnum() or c in (' ', '-', '_', '(', ')', '（', '）', '，', '。')).strip()
            if not filename_base:
                filename_base = f"file_{file_count + 1}"
            
            # 生成文件名
            filename = f"{filename_base}.xlsx"
            output_path = os.path.join(output_dir, filename)
            
            # 如果文件名已存在，添加序号
            counter = 1
            original_filename = filename
            while os.path.exists(output_path):
                name, ext = os.path.splitext(original_filename)
                filename = f"{name}_{counter}{ext}"
                output_path = os.path.join(output_dir, filename)
                counter += 1
            
            # 保存文件
            wb.save(output_path)
            file_count += 1
            
            # 更新进度
            if total_rows > 0:
                progress = file_count / total_rows
                progress_bar.progress(min(progress, 1.0))
                status_text.text(f"已创建 {file_count} 个文件...")
        
        progress_bar.empty()
        status_text.empty()
        
        return file_count
        
    except Exception as e:
        st.error(f"处理文件时出错: {str(e)}")
        raise


def merge_excel_files(excel_files):
    """合并多个Excel文件"""
    try:
        dataframes = []
        
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        # 读取每个Excel文件
        for idx, file_path in enumerate(excel_files, 1):
            try:
                # 读取Excel文件，使用第一行作为列名
                df = pd.read_excel(file_path, header=0)
                
                # 添加源文件名列，用于追踪数据来源
                if '源文件' not in df.columns:
                    df.insert(0, '源文件', os.path.basename(file_path))
                
                dataframes.append(df)
                status_text.text(f"已读取 [{idx}/{len(excel_files)}]: {os.path.basename(file_path)} - {df.shape[0]} 行, {df.shape[1]} 列")
                
                # 更新进度
                progress = idx / len(excel_files)
                progress_bar.progress(progress)
                
            except Exception as e:
                st.warning(f"读取文件失败 {os.path.basename(file_path)}: {str(e)}")
                continue
        
        progress_bar.empty()
        status_text.empty()
        
        if not dataframes:
            raise Exception("没有成功读取任何文件")
        
        # 合并所有数据框
        merged_df = pd.concat(dataframes, ignore_index=True, sort=False)
        
        return merged_df
        
    except Exception as e:
        st.error(f"处理过程中出错: {str(e)}")
        raise


# 页面配置
st.set_page_config(
    page_title="Excel工具",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# 自定义CSS样式（iOS风格）
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: 600;
        color: #000;
        margin-bottom: 0.5rem;
    }
    .sub-header {
        font-size: 1.2rem;
        color: #8e8e93;
        margin-bottom: 2rem;
    }
    .stButton>button {
        width: 100%;
        border-radius: 8px;
        border: none;
        padding: 0.75rem 1.5rem;
        font-size: 1rem;
        font-weight: 500;
    }
    .success-box {
        padding: 1rem;
        border-radius: 8px;
        background-color: #d4edda;
        border: 1px solid #c3e6cb;
        color: #155724;
        margin-top: 1rem;
    }
    .error-box {
        padding: 1rem;
        border-radius: 8px;
        background-color: #f8d7da;
        border: 1px solid #f5c6cb;
        color: #721c24;
        margin-top: 1rem;
    }
</style>
""", unsafe_allow_html=True)

# 标题区域
col1, col2 = st.columns([3, 1])
with col1:
    st.markdown('<div class="main-header">Excel 工具</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-header">拆分与合并 Excel 文件</div>', unsafe_allow_html=True)

# 如果有背景图片，可以在这里显示
# if os.path.exists("cat_bg.png"):
#     st.image("cat_bg.png", use_container_width=False, width=200)

# 功能选择
mode = st.radio(
    "选择功能",
    ["拆分 Excel 文件", "合并 Excel 文件"],
    horizontal=True,
    label_visibility="visible"
)

# 根据模式显示不同的界面
if mode == "拆分 Excel 文件":
    st.markdown("---")
    st.markdown("### 📂 拆分 Excel 文件")
    
    uploaded_file = st.file_uploader(
        "请选择要拆分的 Excel 文件",
        type=['xlsx', 'xls'],
        help="上传一个Excel文件，程序将按行拆分成多个文件"
    )
    
    if uploaded_file is not None:
        # 创建临时文件保存上传的文件
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            tmp_file.write(uploaded_file.getvalue())
            tmp_file_path = tmp_file.name
        
        try:
            # 显示文件信息
            wb = load_workbook(tmp_file_path)
            ws = wb.active
            st.info(f"📄 文件结构: {ws.max_row} 行, {ws.max_column} 列")
            
            if st.button("▶ 开始拆分", type="primary", use_container_width=True):
                with st.spinner("正在拆分文件，请稍候..."):
                    # 创建临时目录保存拆分后的文件
                    with tempfile.TemporaryDirectory() as tmp_dir:
                        try:
                            file_count = split_excel_by_rows(tmp_file_path, tmp_dir)
                            
                            if file_count > 0:
                                # 创建ZIP文件
                                zip_buffer = io.BytesIO()
                                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                                    for root, dirs, files in os.walk(tmp_dir):
                                        for file in files:
                                            file_path = os.path.join(root, file)
                                            arc_name = os.path.relpath(file_path, tmp_dir)
                                            zip_file.write(file_path, arc_name)
                                
                                zip_buffer.seek(0)
                                
                                st.success(f"✅ 拆分完成！共创建了 {file_count} 个文件")
                                
                                # 提供下载按钮
                                st.download_button(
                                    label="📥 下载所有拆分文件 (ZIP)",
                                    data=zip_buffer,
                                    file_name="拆分后的文件.zip",
                                    mime="application/zip",
                                    use_container_width=True
                                )
                            else:
                                st.warning("⚠️ 没有找到需要拆分的数据行")
                                
                        except Exception as e:
                            st.error(f"❌ 拆分过程中出错: {str(e)}")
                            st.exception(e)
        finally:
            # 清理临时文件
            if os.path.exists(tmp_file_path):
                os.unlink(tmp_file_path)

else:
    st.markdown("---")
    st.markdown("### 📥 合并 Excel 文件")
    
    uploaded_files = st.file_uploader(
        "请选择要合并的 Excel 文件（可多选）",
        type=['xlsx', 'xls'],
        accept_multiple_files=True,
        help="可以选择多个Excel文件进行合并"
    )
    
    if uploaded_files and len(uploaded_files) > 0:
        st.info(f"📄 已选择 {len(uploaded_files)} 个文件")
        
        # 显示文件列表
        with st.expander("查看文件列表"):
            for idx, file in enumerate(uploaded_files, 1):
                st.text(f"{idx}. {file.name}")
        
        output_filename = st.text_input(
            "输出文件名",
            value="合并后的Excel.xlsx",
            help="合并后文件的名称"
        )
        
        if st.button("▶ 开始合并", type="primary", use_container_width=True):
            with st.spinner("正在合并文件，请稍候..."):
                try:
                    # 保存上传的文件到临时目录
                    with tempfile.TemporaryDirectory() as tmp_dir:
                        excel_files = []
                        for uploaded_file in uploaded_files:
                            file_path = os.path.join(tmp_dir, uploaded_file.name)
                            with open(file_path, 'wb') as f:
                                f.write(uploaded_file.getbuffer())
                            excel_files.append(file_path)
                        
                        # 合并文件
                        merged_df = merge_excel_files(excel_files)
                        
                        if merged_df is not None and not merged_df.empty:
                            # 保存到临时文件
                            output_path = os.path.join(tmp_dir, output_filename)
                            merged_df.to_excel(output_path, index=False, engine='openpyxl')
                            
                            # 读取文件供下载
                            with open(output_path, 'rb') as f:
                                file_data = f.read()
                            
                            st.success(f"✅ 合并完成！")
                            st.info(f"📊 统计信息: {len(merged_df)} 行, {len(merged_df.columns)} 列")
                            
                            # 提供下载按钮
                            st.download_button(
                                label=f"📥 下载合并后的文件: {output_filename}",
                                data=file_data,
                                file_name=output_filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )
                        else:
                            st.warning("⚠️ 合并后的数据为空")
                            
                except Exception as e:
                    st.error(f"❌ 合并过程中出错: {str(e)}")
                    st.exception(e)

