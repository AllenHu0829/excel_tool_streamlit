import pandas as pd
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import shutil

def split_excel_by_rows(input_file):
    """
    按照表头分割Excel文件，每一行对应一个文件
    表头只有第1行
    文件名按照分割后文件的A2单元格内容命名
    F1~K1需要蓝色填充，L1~M1需要红色填充
    所有列宽根据字符长度自动适应宽度
    """
    try:
        # 使用openpyxl读取原始文件
        source_wb = load_workbook(input_file)
        source_ws = source_wb.active
        
        print(f"Excel文件结构:")
        print(f"最大行数: {source_ws.max_row}")
        print(f"最大列数: {source_ws.max_column}")
        
        # 创建输出目录
        output_dir = os.path.join(os.path.dirname(input_file), "split_files")
        if os.path.exists(output_dir):
            try:
                shutil.rmtree(output_dir)  # 删除旧文件
            except PermissionError:
                print("无法删除旧文件，将覆盖现有文件")
        os.makedirs(output_dir, exist_ok=True)
        
        # 定义颜色填充
        blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # 浅蓝色
        red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")    # 浅红色
        
        # 遍历每一行数据（从第2行开始，因为第1行是表头）
        file_count = 0
        for row_num in range(2, source_ws.max_row + 1):
            # 检查该行是否有数据（检查A列是否有内容）
            if source_ws.cell(row=row_num, column=1).value is None:
                continue
            
            # 创建新的工作簿
            wb = Workbook()
            ws = wb.active
            
            # 复制表头第1行
            for col in range(1, source_ws.max_column + 1):
                source_cell = source_ws.cell(row=1, column=col)
                target_cell = ws.cell(row=1, column=col)
                target_cell.value = source_cell.value
                
                # 应用颜色填充
                if 6 <= col <= 11:  # F1~K1 (列6-11)
                    target_cell.fill = blue_fill
                elif 12 <= col <= 13:  # L1~M1 (列12-13)
                    target_cell.fill = red_fill
            
            # 复制数据行（第2行）
            for col in range(1, source_ws.max_column + 1):
                source_cell = source_ws.cell(row=row_num, column=col)
                target_cell = ws.cell(row=2, column=col)
                target_cell.value = source_cell.value
            
            # 自动调整列宽
            for col in range(1, source_ws.max_column + 1):
                column_letter = ws.cell(row=1, column=col).column_letter
                max_length = 0
                
                # 检查表头和数据行的内容长度
                for row in range(1, 3):  # 检查第1行和第2行
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        # 计算字符长度，中文字符按2个字符计算
                        length = 0
                        for char in str(cell_value):
                            if ord(char) > 127:  # 中文字符
                                length += 2
                            else:
                                length += 1
                        max_length = max(max_length, length)
                
                # 设置列宽，最小宽度为8，最大宽度为50
                adjusted_width = min(max(max_length + 2, 8), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 获取该文件A2单元格的内容作为文件名
            filename_base = str(ws.cell(row=2, column=1).value) if ws.cell(row=2, column=1).value else f"file_{file_count + 1}"
            
            # 清理文件名中的非法字符
            filename_base = "".join(c for c in filename_base if c.isalnum() or c in (' ', '-', '_', '(', ')', '（', '）', '，', '。')).strip()
            if not filename_base:
                filename_base = f"file_{file_count + 1}"
            
            # 生成文件名
            filename = f"{filename_base}.xlsx"
            output_path = os.path.join(output_dir, filename)
            
            # 如果文件名已存在，添加序号
            counter = 1
            original_filename = filename
            while os.path.exists(output_path):
                name, ext = os.path.splitext(original_filename)
                filename = f"{name}_{counter}{ext}"
                output_path = os.path.join(output_dir, filename)
                counter += 1
            
            # 保存文件
            wb.save(output_path)
            print(f"已创建文件: {filename}")
            file_count += 1
        
        print(f"\n分割完成！共创建了 {file_count} 个文件")
        print(f"文件保存在: {output_dir}")
        
    except Exception as e:
        print(f"处理文件时出错: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    input_file = r"c:\Users\AllenHu\excel data\工作簿1.xlsx"
    split_excel_by_rows(input_file)

